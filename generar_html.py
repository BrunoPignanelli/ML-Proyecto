"""
Genera petinsa_envios.html — app standalone para PETINSA.
Ejecutar: python generar_html.py
"""
import openpyxl, re, unicodedata, json
from pathlib import Path
import copy

# ═══════════════════════════════════════════════════════
# 1. FUNCIONES DE CLASIFICACIÓN
# ═══════════════════════════════════════════════════════

def norm(s):
    if s is None: return ''
    s = str(s).replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def extract_rim(desc):
    d = str(desc).upper()
    m = re.search(r'\d\s+-\s+(\d{2,3})\b', d)
    if m: return float(m.group(1))
    m = re.search(r'-(\d{2,3})\b', d)
    if m: return float(m.group(1))
    m = re.search(r'R\s*(\d{2}(?:\.\d)?)\b', d)
    if m: return float(m.group(1))
    return None

def extract_amp(desc):
    m = re.search(r'\b(\d{2,3})A\b', str(desc).upper())
    return int(m.group(1)) if m else None

UNIFIED_CATS = {
    'cub_moto':          'Cubierta Moto',
    'cub_auto_r12_r14':  'Cubierta Auto R12-R14',
    'cub_auto_r15_r19':  'Cubierta Auto/Camioneta R15-R19',
    'cub_camion_r20_r22':'Cubierta Camion R20-R22.5',
    'cub_agro_del':      'Cubierta Agricola Delantera',
    'cub_agro_tras_med': 'Cubierta Agricola Trasera Med. 24-26"',
    'cub_agro_tras_gde': 'Cubierta Agricola Trasera Gde. 28-36"',
    'cub_agro_tras_xgde':'Cubierta Agricola/Vial +38"',
    'camara':            'Camara',
    'bateria_chica':     'Bateria Chica (<=110 Ah)',
    'bateria_grande':    'Bateria Grande (>110 Ah)',
    'lubricante_caja':   'Lubricante en caja/lata',
    'lubricante_tambor': 'Lubricante Tambor 205L',
    'bulto_general':     'Bulto General',
}

def classify(familia, desc):
    f = str(familia).upper()
    d = str(desc).upper()
    if 'LUBRICANTE' in f:
        return 'lubricante_tambor' if ('TAMBOR' in d or '205' in d) else 'lubricante_caja'
    if 'BATERIA' in f:
        amp = extract_amp(desc)
        return 'bateria_grande' if amp and amp > 110 else 'bateria_chica'
    if 'ACCES' in f or 'CAM MOTO' in f:
        return 'camara'
    if 'MOTO' in f and 'CAM' not in f:
        return 'cub_moto'
    if 'GIGANTE' in f:
        return 'cub_camion_r20_r22'
    if 'VIAL' in f or 'IND.' in f:
        rim = extract_rim(desc)
        return 'cub_agro_tras_xgde' if rim and rim >= 38 else 'cub_agro_tras_gde'
    if 'AGR DEL' in f or 'AGRICOLA DEL' in f:
        rim = extract_rim(desc)
        if rim:
            if rim >= 38: return 'cub_agro_tras_xgde'
            if rim >= 28: return 'cub_agro_tras_gde'
            if rim >= 24: return 'cub_agro_tras_med'
        return 'cub_agro_del'
    if 'AGR TRAS' in f or 'AGRICOLA TRAS' in f:
        rim = extract_rim(desc)
        if rim:
            if rim >= 38: return 'cub_agro_tras_xgde'
            if rim >= 30: return 'cub_agro_tras_gde'
            if rim >= 24: return 'cub_agro_tras_med'
        return 'cub_agro_tras_med'
    if 'PASEO' in f:
        rim = extract_rim(desc)
        return 'cub_auto_r12_r14' if rim and rim <= 14 else 'cub_auto_r15_r19'
    if 'CAMIONETA' in f or 'PICK UP' in f:
        rim = extract_rim(desc)
        if rim:
            if rim <= 14: return 'cub_auto_r12_r14'
            if rim <= 19: return 'cub_auto_r15_r19'
            return 'cub_camion_r20_r22'
        return 'cub_auto_r15_r19'
    return 'bulto_general'

# ═══════════════════════════════════════════════════════
# 2. MAPEO: categoría unificada → label de la agencia
# ═══════════════════════════════════════════════════════

MAPPING = {
    'DAC': {
        'cub_moto':'Atado moto','cub_auto_r12_r14':'Rodado 12,13,14',
        'cub_auto_r15_r19':'Rodado 15 a 18','cub_camion_r20_r22':'Rodado 20 a 22,5',
        'cub_agro_del':'Cubierta tractor AGRO 24 a 42','cub_agro_tras_med':'Cubierta tractor AGRO 24 a 42',
        'cub_agro_tras_gde':'Cubierta tractor AGRO 24 a 42','cub_agro_tras_xgde':'Cubierta tractor AGRO 24 a 42',
        'camara':'Bulto hasta 15kg','bateria_chica':'Bateria chica','bateria_grande':'Bateria grande',
        'lubricante_caja':'Bulto hasta 15kg','lubricante_tambor':'Tambor 205lts','bulto_general':'Bulto hasta 15kg',
    },
    'NASAZZI': {
        'cub_moto':'Rodado 15 a 19','cub_auto_r12_r14':'Rodado 13 y 14',
        'cub_auto_r15_r19':'Rodado 15 a 19','cub_camion_r20_r22':'Rodado 20 a 22,5',
        'cub_agro_del':'Rodado 24 y 25','cub_agro_tras_med':'Rodado 24 y 25',
        'cub_agro_tras_gde':'Rodado 26 a 32','cub_agro_tras_xgde':'Cub viales de 38 a 46 extra grande',
        'camara':'Bultos','bateria_chica':'Bat hasta 110 amp','bateria_grande':'Bat mayor 110 amp',
        'lubricante_caja':'Bultos lubricantes','lubricante_tambor':'Tambor 205lts','bulto_general':'Bultos',
    },
    'MEGAM': {
        'cub_moto':'Bat y bultos','cub_auto_r12_r14':'Cub 13 y 14',
        'cub_auto_r15_r19':'Cubierta auto (15 a 17)','cub_camion_r20_r22':'Cubierta camion',
        'cub_agro_del':'Tractor chico','cub_agro_tras_med':'Tractor chico',
        'cub_agro_tras_gde':'Tractor grande','cub_agro_tras_xgde':'Tractor grande',
        'camara':'Bat y bultos','bateria_chica':'Bat y bultos','bateria_grande':'Bat y bultos',
        'lubricante_caja':'Bat y bultos','lubricante_tambor':'Tambor 205lts','bulto_general':'Bat y bultos',
    },
    'Transportes Nagar sas (Expreso Ruta 1)': {
        'cub_moto':'Auto','cub_auto_r12_r14':'Auto','cub_auto_r15_r19':'Camioneta',
        'cub_camion_r20_r22':'Camion','cub_agro_del':'Tractor del y 17,5',
        'cub_agro_tras_med':'Tractor traseras','cub_agro_tras_gde':'Tractor grande',
        'cub_agro_tras_xgde':'Tractor grande','camara':'Bultos','bateria_chica':'Baterias',
        'bateria_grande':'Baterias','lubricante_caja':'Bultos','lubricante_tambor':'Tambor','bulto_general':'Bultos',
    },
    'SELEGUIN': {
        'cub_moto':'Bultos, atados, bolsas, cajas','cub_auto_r12_r14':'Cubierta auto',
        'cub_auto_r15_r19':'Cubierta camioneta','cub_camion_r20_r22':'Cubierta camion',
        'cub_agro_del':'Bultos, atados, bolsas, cajas','cub_agro_tras_med':'Tractor mediana 24 a 26',
        'cub_agro_tras_gde':'Tractor grande 28 a 30','cub_agro_tras_xgde':'Tractor extra grande +30',
        'camara':'Bultos, atados, bolsas, cajas','bateria_chica':'Baterias chicas','bateria_grande':'Baterias grandes',
        'lubricante_caja':'Bultos, atados, bolsas, cajas','lubricante_tambor':None,'bulto_general':'Bultos, atados, bolsas, cajas',
    },
    'EXPRESO ROCHA': {
        'cub_moto':'Bultos camaras','cub_auto_r12_r14':'Cubierta chica hasta 205/70.16',
        'cub_auto_r15_r19':'Cubierta grande hasta 750.16','cub_camion_r20_r22':'Cubierta camion',
        'cub_agro_del':'Neumatico agricolas y viales hasta 18.4/15-26',
        'cub_agro_tras_med':'Neum. Agricolas y viales mas de 23.1/18-26',
        'cub_agro_tras_gde':'Neum. Agricolas y viales mas de 23.1/18-26',
        'cub_agro_tras_xgde':'Neum. Agricolas y viales mas de 23.1/18-26',
        'camara':'Bultos camaras','bateria_chica':'Baterias','bateria_grande':'Baterias',
        'lubricante_caja':'Bultos camaras','lubricante_tambor':None,'bulto_general':'Bultos camaras',
    },
    'BULEVAR (ACC)': {
        'cub_moto':'Atado cub moto','cub_auto_r12_r14':'Cub auto  hasta rod 14',
        'cub_auto_r15_r19':'Cub camioneta 15 a 19','cub_camion_r20_r22':'Cub camion',
        'cub_agro_del':'Cub tractor chica (hasta 22)','cub_agro_tras_med':'Cub tractor mediana (24)',
        'cub_agro_tras_gde':'Cub tractor grande (de 25 a 36)','cub_agro_tras_xgde':'Cub vial rod 38 a 42',
        'camara':'Bolsas','bateria_chica':'Bat chicas','bateria_grande':'Bat grandes',
        'lubricante_caja':'Caja aceite 1 y 4lts','lubricante_tambor':'Tambor 205lts','bulto_general':'Bolsas',
    },
    'PERICO': {
        'cub_moto':'Atado cub moto','cub_auto_r12_r14':'Cub auto  hasta rod 14',
        'cub_auto_r15_r19':'Cub camioneta 15 a 19','cub_camion_r20_r22':'Cub camion',
        'cub_agro_del':'Cub tractor chica','cub_agro_tras_med':'Cub tractor mediana',
        'cub_agro_tras_gde':'Cub tractor grande','cub_agro_tras_xgde':'Cub vial rod 38 a 42',
        'camara':'Bolsas','bateria_chica':'Bat chicas','bateria_grande':'Bat grandes',
        'lubricante_caja':'Caja aceite 1 y 4lts','lubricante_tambor':'Tambor 205lts','bulto_general':'Bolsas',
    },
    'TRUJILLO': {
        'cub_moto':'Bultos','cub_auto_r12_r14':'Paseo hasta 16','cub_auto_r15_r19':'Paseo hasta 16',
        'cub_camion_r20_r22':'Camion','cub_agro_del':'Agricola delantera',
        'cub_agro_tras_med':'Agricola mediana','cub_agro_tras_gde':'Agricola extra grande',
        'cub_agro_tras_xgde':'Agricola extra grande','camara':'Bultos',
        'bateria_chica':'Bat hasta 150 amp','bateria_grande':'Bat + 150 amp',
        'lubricante_caja':'Cajas lubricantes','lubricante_tambor':'Tambor 205lts','bulto_general':'Bultos',
    },
    'GONFER': {
        'cub_moto':'Bultos','cub_auto_r12_r14':'Cubierta de auto',
        'cub_auto_r15_r19':'Cubierta de camioneta','cub_camion_r20_r22':'Cubierta de camion chico',
        'cub_agro_del':'Agricola chica','cub_agro_tras_med':'Agricola hasta rod 28',
        'cub_agro_tras_gde':'Agricola grande rod 30 en adelante','cub_agro_tras_xgde':'Agricola grande rod 30 en adelante',
        'camara':'Bultos','bateria_chica':'Bat hasta 110 amp','bateria_grande':'Bat mayor 110 amp',
        'lubricante_caja':'Bultos lubricantes','lubricante_tambor':'Tambor 205lts','bulto_general':'Bultos',
    },
    '3EME (El Chambon)': {
        'cub_moto':'Atado cub moto','cub_auto_r12_r14':'Cub auto  hasta rod 14',
        'cub_auto_r15_r19':'Cub camioneta 15 a 19','cub_camion_r20_r22':'Cub camion',
        'cub_agro_del':'Cub tractor chica','cub_agro_tras_med':'Cub tractor mediana',
        'cub_agro_tras_gde':'Cub tractor grande','cub_agro_tras_xgde':'Cub vial rod 38 a 42',
        'camara':'Bolsas','bateria_chica':'Bat chicas','bateria_grande':'Bat grandes',
        'lubricante_caja':'Caja aceite 1 y 4lts','lubricante_tambor':'Tambor 205lts','bulto_general':'Bolsas',
    },
    'Martin Escudero (Lascano)': {
        'cub_moto':'Auto','cub_auto_r12_r14':'Auto','cub_auto_r15_r19':'Camioneta',
        'cub_camion_r20_r22':'Camion','cub_agro_del':'Agricola chica',
        'cub_agro_tras_med':'Agricola mediana','cub_agro_tras_gde':'Agricola grande',
        'cub_agro_tras_xgde':'Agricola grande','camara':'Bultos/atados',
        'bateria_chica':'Baterias chicas','bateria_grande':'Baterias grandes',
        'lubricante_caja':'Bultos lubricantes','lubricante_tambor':'Tambor aceite','bulto_general':'Bultos/atados',
    },
    'Arzuaga': {
        'cub_moto':'Atado cubiertas moto','cub_auto_r12_r14':'cubiertas r13 r14',
        'cub_auto_r15_r19':'cubiertas r16 r18','cub_camion_r20_r22':'Cubiertas camion',
        'cub_agro_del':'Cubiertas tractor chica','cub_agro_tras_med':'Cubiertas tractor mediana',
        'cub_agro_tras_gde':'Cubiertas tractor grande','cub_agro_tras_xgde':'Cubiertas tractor grande',
        'camara':'Paquetes/cajas','bateria_chica':'Baterias chicas','bateria_grande':'Baterias camion',
        'lubricante_caja':'Paquetes/cajas','lubricante_tambor':'Tanque aceite','bulto_general':'Paquetes/cajas',
    },
    'Franchi': {
        'cub_moto':'Bultos','cub_auto_r12_r14':'Cub.auto','cub_auto_r15_r19':'Cub.camioneta',
        'cub_camion_r20_r22':'Cub.camion','cub_agro_del':'Bultos',
        'cub_agro_tras_med':'Cub. agricola mediana','cub_agro_tras_gde':'Cub. agricola gde',
        'cub_agro_tras_xgde':'Cub. agricola gde','camara':'Bultos',
        'bateria_chica':'Baterias chicas','bateria_grande':'Baterias grandes',
        'lubricante_caja':'Bultos','lubricante_tambor':None,'bulto_general':'Bultos',
    },
}

# ═══════════════════════════════════════════════════════
# 3. CARGA DE DATOS
# ═══════════════════════════════════════════════════════

print("Cargando Libro1.xlsx...")
wb1 = openpyxl.load_workbook('Libro1.xlsx')
ws1 = wb1['Hoja1']
rows1 = list(ws1.iter_rows(values_only=True))
cols1 = [str(c).strip() if c else '' for c in rows1[0]]

catalog = []
for row in rows1[1:]:
    if not any(row): continue
    r = dict(zip(cols1, row))
    fam  = str(r.get('Familia', '') or '').strip()
    desc = str(r.get('Descripción de Producto', '') or '').strip()
    cod_raw = r.get('Prod.')
    cod = str(cod_raw).strip().rstrip('.0') if cod_raw is not None else ''
    if not cod or not desc: continue
    ucat = classify(fam, desc)
    def safe_int(v):
        try: return int(float(v or 0))
        except: return 0
    def safe_float(v):
        try: return round(float(v or 0), 2)
        except: return 0.0
    catalog.append({
        'c': cod, 'd': desc, 'f': fam,
        'r': str(r.get('Ramo', '') or '').strip(),
        'pl': safe_float(r.get('P. Lista')),
        'st': safe_int(r.get('Stock')),
        'pe': safe_int(r.get('Pedido')),
        'd3': safe_int(r.get('Disp 30d')),
        'ms': safe_float(r.get('Meses Stk')),
        'uc': ucat,
        'cn': UNIFIED_CATS[ucat],
    })

print(f"  {len(catalog)} productos cargados.")

print("Cargando COMPARATIVA AGENCIAS.xlsx (Mayo 2026)...")
wb2 = openpyxl.load_workbook('COMPARATIVA AGENCIAS.xlsx', data_only=True)
ws2 = wb2['Costo ag. Mayo 2026']
ag_rows = list(ws2.iter_rows(values_only=True))
agency_name_row = ag_rows[5]
canje_row       = ag_rows[2]
freq_row        = ag_rows[4]
data_rows       = ag_rows[7:]

ncols = len(agency_name_row)
acg = {}
col = 0
while col < ncols:
    name = agency_name_row[col]
    if name and str(name).strip():
        nxt = col + 1
        while nxt < ncols and (agency_name_row[nxt] is None or str(agency_name_row[nxt]).strip() == ''):
            nxt += 1
        acg[str(name).strip()] = list(range(col, nxt))
        col = nxt
    else:
        col += 1

IVA = 1.22

def pick_price_col(ag, ci):
    pc = ci[1] if len(ci) > 1 else ci[0]
    if 'GONFER' in ag.upper() and len(ci) >= 4: pc = ci[3]
    return ci[0], pc

def parse_canje(val):
    if isinstance(val, (int, float)): return min(float(val), 1.0)
    return 0.0

def parse_freq(text):
    if not text: return ('L a V', 5)
    t = str(text).lower()
    if 'cuando tiene carga' in t: return ('Bajo pedido', 1)
    if 'mart y jue' in t or 'martes y jueves' in t: return ('Mar y Jue', 2)
    if 'l, mier y v' in t: return ('L, Mier y V', 3)
    return ('L a V', 5)

agencies_meta  = {}
agency_prices  = {}

for ag, ci in acg.items():
    cat_col, price_col = pick_price_col(ag, ci)
    prices = {}
    for row in data_rows:
        cat   = row[cat_col]   if cat_col   < len(row) else None
        price = row[price_col] if price_col < len(row) else None
        if cat and isinstance(cat, str) and cat.strip() and price and isinstance(price, (int, float)):
            prices[norm(cat.strip())] = float(price)
    agency_prices[ag] = prices
    canje_col = ci[2] if ('GONFER' in ag.upper() and len(ci) >= 3) else price_col
    raw_canje = canje_row[canje_col] if canje_col < len(canje_row) else None
    raw_freq  = freq_row[cat_col]    if cat_col   < len(freq_row)  else None
    freq_label, freq_days = parse_freq(raw_freq)
    agencies_meta[ag] = {
        'nombre': ag,
        'canje':  parse_canje(raw_canje),
        'freq':   freq_label,
        'dias':   freq_days,
        'iva':    'sin IVA' if 'GONFER' in ag.upper() else 'IVA incl.',
    }

def build_prices(mapping):
    pj = {}
    for ag, ag_map in mapping.items():
        if ag not in agencies_meta: continue
        pj[ag] = {}
        for ucat, label in ag_map.items():
            if label is None: continue
            bruto = agency_prices.get(ag, {}).get(norm(label))
            if bruto:
                if agencies_meta[ag]['iva'] == 'sin IVA':
                    bruto = round(bruto * IVA, 2)
                pj[ag][ucat] = bruto
    return pj

prices_data = build_prices(MAPPING)
print(f"  {len(agencies_meta)} agencias procesadas.")

# ═══════════════════════════════════════════════════════
# 4. SERIALIZAR JSON
# ═══════════════════════════════════════════════════════
catalog_json   = json.dumps(catalog,                      ensure_ascii=False, separators=(',',':'))
agencies_json  = json.dumps(list(agencies_meta.values()), ensure_ascii=False, separators=(',',':'))
prices_json    = json.dumps(prices_data,                  ensure_ascii=False, separators=(',',':'))
ucats_json     = json.dumps(UNIFIED_CATS,                 ensure_ascii=False, separators=(',',':'))
ag_names_json  = json.dumps(list(MAPPING.keys()),         ensure_ascii=False, separators=(',',':'))

# ═══════════════════════════════════════════════════════
# 5. HTML TEMPLATE
#    Usamos string normal + replace() para evitar
#    problemas de escaping con f-strings y JS
# ═══════════════════════════════════════════════════════

HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>PETINSA - Sistema de Envios</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css" rel="stylesheet">
<style>
  :root { --azul: #1a3a5c; --azul-claro: #e8f0f8; }
  body  { background: #f4f6f9; font-size: .91rem; }
  .navbar  { background: var(--azul) !important; }
  .tab-btn { cursor:pointer; padding:.45rem 1.1rem; border:none; background:transparent;
             border-bottom:3px solid transparent; color:#555; font-weight:500; }
  .tab-btn.active { border-bottom-color:var(--azul); color:var(--azul); font-weight:700; }
  .tab-btn:hover:not(.active) { background:#f0f0f0; border-radius:4px 4px 0 0; }
  .card  { border:none; box-shadow:0 1px 6px rgba(0,0,0,.09); border-radius:10px; }
  .ch    { background:var(--azul); color:#fff; border-radius:10px 10px 0 0; padding:.65rem 1rem; font-weight:600; }
  .btn-p { background:var(--azul); color:#fff; border:none; }
  .btn-p:hover { background:#142d47; color:#fff; }
  #ac-box { position:absolute; z-index:9999; background:#fff; border:1px solid #ccc;
            border-radius:0 0 8px 8px; width:100%; max-height:270px; overflow-y:auto;
            box-shadow:0 4px 14px rgba(0,0,0,.18); top:100%; left:0; }
  .ac-it  { padding:.42rem .8rem; cursor:pointer; border-bottom:1px solid #f0f0f0; }
  .ac-it:hover { background:var(--azul-claro); }
  .ac-cod { font-weight:700; color:var(--azul); margin-right:.4rem; font-size:.82rem; }
  .ac-fam { font-size:.76rem; color:#888; }
  .bcat   { background:var(--azul-claro); color:var(--azul); border-radius:20px;
            padding:1px 8px; font-size:.74rem; }
  .rank-1 { background:#fff9e6; }
  .rank-2 { background:#f0fff4; }
  .rank-3 { background:#f0f8ff; }
  .neto0  { color:#198754; font-weight:700; }
  .ptag   { font-weight:600; color:var(--azul); }
  .qty-inp{ width:68px !important; text-align:center; }
  .cat-tr:hover { background:var(--azul-claro); cursor:pointer; }
  .pg-btn { cursor:pointer; padding:3px 9px; border-radius:5px;
            border:1px solid #ddd; margin:0 2px; background:#fff; font-size:.84rem; }
  .pg-btn.active { background:var(--azul); color:#fff; border-color:var(--azul); }
  #ocnt   { background:var(--azul); color:#fff; border-radius:50%;
            padding:0 5px; font-size:.72rem; margin-left:3px; }
  .field-row label { font-weight:600; font-size:.84rem; color:#444; }
</style>
</head>
<body>

<nav class="navbar navbar-dark px-3 py-2 mb-0">
  <span class="navbar-brand fw-bold fs-5">
    <i class="bi bi-truck me-2"></i>PETINSA &mdash; Sistema de Envios
  </span>
  <span class="text-white-50 small">Tarifas Mayo 2026</span>
</nav>

<div class="container-fluid px-3 pt-3">

  <!-- TABS -->
  <div class="d-flex border-bottom mb-3">
    <button class="tab-btn active" id="btn-calc" onclick="showTab('calc')">
      <i class="bi bi-calculator me-1"></i>Calcular Envio
      <span id="ocnt" style="display:none">0</span>
    </button>
    <button class="tab-btn" id="btn-cat" onclick="showTab('cat')">
      <i class="bi bi-grid me-1"></i>Catalogo de Productos
    </button>
  </div>

  <!-- ====== TAB: CALCULADORA ====== -->
  <div id="tab-calc">
    <div class="row g-3">

      <!-- Datos del pedido -->
      <div class="col-12">
        <div class="card">
          <div class="ch"><i class="bi bi-file-text me-2"></i>Datos del pedido</div>
          <div class="card-body">
            <div class="row g-2">
              <div class="col-md-2 col-sm-4">
                <label class="form-label mb-1">N° Pedido</label>
                <input id="f-nped" class="form-control form-control-sm" placeholder="Ej: 00123">
              </div>
              <div class="col-md-3 col-sm-8">
                <label class="form-label mb-1">Cliente</label>
                <input id="f-cliente" class="form-control form-control-sm" placeholder="Nombre del cliente">
              </div>
              <div class="col-md-3 col-sm-6">
                <label class="form-label mb-1">Localidad / Destino</label>
                <input id="f-dest" class="form-control form-control-sm" placeholder="Ciudad de destino">
              </div>
              <div class="col-md-2 col-sm-3">
                <label class="form-label mb-1">Fecha</label>
                <input id="f-fecha" type="date" class="form-control form-control-sm">
              </div>
              <div class="col-md-2 col-sm-3">
                <label class="form-label mb-1">Vendedor</label>
                <input id="f-vendedor" class="form-control form-control-sm" placeholder="Nombre">
              </div>
              <div class="col-12">
                <label class="form-label mb-1">Observaciones</label>
                <input id="f-obs" class="form-control form-control-sm" placeholder="Opcional">
              </div>
            </div>
          </div>
        </div>
      </div>

      <!-- Buscador de productos -->
      <div class="col-12">
        <div class="card">
          <div class="ch"><i class="bi bi-search me-2"></i>Agregar productos al pedido</div>
          <div class="card-body pb-2">
            <div class="position-relative">
              <input id="prod-search" class="form-control"
                     placeholder="Buscar por codigo, descripcion o ramo..."
                     autocomplete="off">
              <div id="ac-box" style="display:none"></div>
            </div>
            <div class="text-muted small mt-1">Escribi al menos 2 caracteres para buscar</div>
          </div>
        </div>
      </div>

      <!-- Tabla del pedido -->
      <div class="col-12">
        <div class="card">
          <div class="ch d-flex justify-content-between align-items-center">
            <span><i class="bi bi-cart3 me-2"></i>Lineas del pedido</span>
            <div>
              <button class="btn btn-sm btn-light me-2" onclick="clearOrder()">
                <i class="bi bi-x-circle me-1"></i>Limpiar
              </button>
              <button class="btn btn-sm btn-warning fw-bold" onclick="calculate()">
                <i class="bi bi-calculator me-1"></i>Calcular envio
              </button>
            </div>
          </div>
          <div class="card-body p-0">
            <div id="order-empty" class="text-center text-muted py-4">
              <i class="bi bi-cart3 fs-2 d-block mb-2 opacity-30"></i>
              Busca productos arriba para armar el pedido
            </div>
            <table class="table table-sm mb-0" id="order-table" style="display:none">
              <thead class="table-light">
                <tr>
                  <th>Codigo</th><th>Descripcion</th><th>Categoria envio</th>
                  <th class="text-center" style="width:90px">Cantidad</th>
                  <th style="width:40px"></th>
                </tr>
              </thead>
              <tbody id="order-body"></tbody>
            </table>
          </div>
        </div>
      </div>

      <!-- RESULTADOS -->
      <div class="col-12" id="results-section" style="display:none">

        <!-- Interpretacion del canje -->
        <div class="card mb-3">
          <div class="card-body py-2">
            <div class="d-flex align-items-center gap-4 flex-wrap">
              <span class="fw-semibold"><i class="bi bi-percent me-1"></i>Canje:</span>
              <div class="form-check form-check-inline mb-0">
                <input class="form-check-input" type="radio" name="cj" id="cj-a" value="A" checked>
                <label class="form-check-label" for="cj-a">
                  A &mdash; Canje = ahorro real en efectivo (neto puede ser $0)
                </label>
              </div>
              <div class="form-check form-check-inline mb-0">
                <input class="form-check-input" type="radio" name="cj" id="cj-b" value="B">
                <label class="form-check-label" for="cj-b">
                  B &mdash; Canje tiene costo de oportunidad
                </label>
              </div>
              <div id="margin-wrap" style="display:none" class="d-flex align-items-center gap-2">
                <label class="mb-0 small">Costo/Venta goma:</label>
                <input type="range" min="0" max="100" value="60" id="margin-sl"
                       class="form-range" style="width:100px">
                <span id="margin-lbl">60%</span>
              </div>
            </div>
          </div>
        </div>

        <!-- Cards resumen -->
        <div class="row g-3 mb-3" id="summary-cards"></div>

        <!-- Modo 1 -->
        <div class="card mb-3">
          <div class="ch"><i class="bi bi-building me-2"></i>Modo 1 &mdash; Todo en una sola agencia</div>
          <div class="card-body p-0">
            <div class="table-responsive">
              <table class="table table-sm table-hover mb-0">
                <thead class="table-light">
                  <tr>
                    <th>#</th><th>Agencia</th>
                    <th class="text-end">Total bruto</th>
                    <th class="text-end">Total neto</th>
                    <th>Canje</th><th>Frecuencia</th><th>Dias/sem</th>
                  </tr>
                </thead>
                <tbody id="m1-body"></tbody>
              </table>
            </div>
          </div>
        </div>

        <!-- Modo 2 -->
        <div class="card">
          <div class="ch"><i class="bi bi-diagram-3 me-2"></i>Modo 2 &mdash; Mejor agencia por producto</div>
          <div class="card-body p-0">
            <div class="table-responsive">
              <table class="table table-sm table-hover mb-0">
                <thead class="table-light">
                  <tr>
                    <th>Codigo</th><th>Descripcion</th><th>Categoria</th>
                    <th class="text-center">Cant.</th>
                    <th>Mejor Agencia</th>
                    <th class="text-end">$/u. neto</th>
                    <th class="text-end">Subtotal</th>
                    <th>Frecuencia</th>
                  </tr>
                </thead>
                <tbody id="m2-body"></tbody>
                <tfoot>
                  <tr class="table-light fw-bold">
                    <td colspan="6" class="text-end">Total Modo 2:</td>
                    <td class="text-end ptag" id="m2-total">-</td>
                    <td></td>
                  </tr>
                </tfoot>
              </table>
            </div>
          </div>
        </div>

      </div><!-- /results-section -->
    </div><!-- /row -->
  </div><!-- /tab-calc -->

  <!-- ====== TAB: CATALOGO ====== -->
  <div id="tab-cat" style="display:none">
    <div class="card">
      <div class="ch"><i class="bi bi-grid me-2"></i>Catalogo de Productos (1437 articulos)</div>
      <div class="card-body">
        <div class="row g-2 mb-3">
          <div class="col-md-5">
            <input id="cat-q" class="form-control" placeholder="Buscar codigo, descripcion, ramo...">
          </div>
          <div class="col-md-3">
            <select id="cat-fam" class="form-select">
              <option value="">Todas las familias</option>
            </select>
          </div>
          <div class="col-md-4 text-end text-muted small d-flex align-items-center justify-content-end">
            <span id="cat-cnt"></span>
          </div>
        </div>
        <div class="table-responsive">
          <table class="table table-sm table-hover mb-0">
            <thead class="table-light">
              <tr>
                <th>Codigo</th><th>Descripcion</th><th>Familia</th><th>Ramo</th>
                <th class="text-end">P. Lista</th>
                <th class="text-end">Stock</th>
                <th class="text-end">Disp. 30d</th>
                <th class="text-end">Meses Stk</th>
                <th>Cat. Envio</th>
                <th></th>
              </tr>
            </thead>
            <tbody id="cat-body"></tbody>
          </table>
        </div>
        <div class="d-flex justify-content-between align-items-center mt-3">
          <span class="text-muted small" id="cat-pg-info"></span>
          <div id="cat-pg"></div>
        </div>
      </div>
    </div>
  </div><!-- /tab-cat -->

</div><!-- /container -->

<!-- DATOS -->
<script>
const CATALOG   = __CATALOG__;
const AGENCIES  = __AGENCIES__;
const PRICES    = __PRICES__;
const UCATS     = __UCATS__;
const AG_NAMES  = __AG_NAMES__;
</script>

<!-- LOGICA -->
<script>
// ── Estado ─────────────────────────────────────────────────────────────────
const order = new Map();   // cod -> {...prod, qty}
let catFiltered = [];
let catPage = 1;
const PG = 25;

// ── Utilidades ──────────────────────────────────────────────────────────────
const $ = id => document.getElementById(id);
const fmt = n => (n === null || n === undefined || isNaN(n))
  ? '-' : '$' + Math.round(n).toLocaleString('es-UY');
const today = () => new Date().toISOString().slice(0,10);

function canjeF(cj) {
  const mode = document.querySelector('input[name="cj"]:checked').value;
  if (mode === 'A') return 1 - cj;
  const mg = parseFloat($('margin-sl').value) / 100;
  return 1 - cj * (1 - mg);
}

// ── Tabs ────────────────────────────────────────────────────────────────────
function showTab(t) {
  ['calc','cat'].forEach(id => {
    $('tab-'+id).style.display = (t === id) ? '' : 'none';
    $('btn-'+id).classList.toggle('active', t === id);
  });
}

// ── Autocomplete ────────────────────────────────────────────────────────────
let acResults = [];
let acIdx = -1;

function acSearch(q) {
  if (q.length < 2) { $('ac-box').style.display = 'none'; acResults = []; return; }
  const ql = q.toLowerCase();
  acResults = CATALOG.filter(p =>
    p.c.toLowerCase().includes(ql) ||
    p.d.toLowerCase().includes(ql) ||
    p.r.toLowerCase().includes(ql)
  ).slice(0, 14);
  renderAC();
}

function renderAC() {
  const box = $('ac-box');
  if (!acResults.length) { box.style.display = 'none'; return; }
  box.innerHTML = acResults.map((p, i) =>
    '<div class="ac-it" data-i="' + i + '">' +
    '<span class="ac-cod">' + esc(p.c) + '</span>' + esc(p.d.substring(0,65)) +
    '<br><span class="ac-fam">' + esc(p.f) + ' &middot; ' + esc(p.cn) + '</span>' +
    '</div>'
  ).join('');
  box.querySelectorAll('.ac-it').forEach(el => {
    el.addEventListener('mousedown', e => {
      e.preventDefault();
      addProduct(acResults[parseInt(el.dataset.i)]);
    });
  });
  box.style.display = '';
  acIdx = -1;
}

function acKey(e) {
  if (e.key === 'ArrowDown') { acIdx = Math.min(acIdx+1, acResults.length-1); hlAC(); e.preventDefault(); }
  else if (e.key === 'ArrowUp')  { acIdx = Math.max(acIdx-1, 0); hlAC(); e.preventDefault(); }
  else if (e.key === 'Enter' && acIdx >= 0) { addProduct(acResults[acIdx]); e.preventDefault(); }
  else if (e.key === 'Escape') { $('ac-box').style.display='none'; }
}

function hlAC() {
  $('ac-box').querySelectorAll('.ac-it').forEach((el,i) =>
    el.classList.toggle('bg-info', i === acIdx));
}

function esc(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

// setup search listeners
const searchEl = $('prod-search');
searchEl.addEventListener('input', () => acSearch(searchEl.value.trim()));
searchEl.addEventListener('keydown', acKey);
document.addEventListener('click', e => {
  if (!e.target.closest('#prod-search') && !e.target.closest('#ac-box'))
    $('ac-box').style.display = 'none';
});

// ── Pedido ──────────────────────────────────────────────────────────────────
function addProduct(p) {
  $('ac-box').style.display = 'none';
  searchEl.value = '';
  acResults = [];
  if (order.has(p.c)) {
    order.get(p.c).qty += 1;
  } else {
    order.set(p.c, Object.assign({}, p, {qty: 1}));
  }
  renderOrder();
  $('results-section').style.display = 'none';
}

function removeItem(cod) {
  order.delete(cod);
  renderOrder();
  $('results-section').style.display = 'none';
}

function updateQty(cod, val) {
  const q = parseInt(val);
  if (isNaN(q) || q <= 0) { removeItem(cod); return; }
  if (order.has(cod)) order.get(cod).qty = q;
}

function clearOrder() {
  order.clear();
  renderOrder();
  $('results-section').style.display = 'none';
}

function renderOrder() {
  const n = order.size;
  const badge = $('ocnt');
  badge.textContent = n;
  badge.style.display = n > 0 ? '' : 'none';
  $('order-empty').style.display  = n === 0 ? '' : 'none';
  $('order-table').style.display  = n > 0   ? '' : 'none';
  $('order-body').innerHTML = [...order.values()].map(p =>
    '<tr>' +
    '<td><code>' + esc(p.c) + '</code></td>' +
    '<td>' + esc(p.d.substring(0,55)) + '</td>' +
    '<td><span class="bcat">' + esc(p.cn) + '</span></td>' +
    '<td class="text-center">' +
      '<input type="number" class="form-control form-control-sm qty-inp d-inline-block"' +
      ' value="' + p.qty + '" min="1"' +
      ' data-cod="' + esc(p.c) + '">' +
    '</td>' +
    '<td><button class="btn btn-sm btn-outline-danger py-0" data-cod="' + esc(p.c) + '">' +
      '<i class="bi bi-trash3"></i></button></td>' +
    '</tr>'
  ).join('');

  // Attach listeners
  $('order-body').querySelectorAll('input[data-cod]').forEach(el => {
    el.addEventListener('change', () => updateQty(el.dataset.cod, el.value));
  });
  $('order-body').querySelectorAll('button[data-cod]').forEach(el => {
    el.addEventListener('click', () => removeItem(el.dataset.cod));
  });
}

// ── Calcular ─────────────────────────────────────────────────────────────────
function getBruto(agNom, ucat) {
  return ((PRICES[agNom] || {})[ucat]) ?? null;
}
function getNeto(agNom, ucat) {
  const b = getBruto(agNom, ucat);
  if (b === null) return null;
  const ag = AGENCIES.find(a => a.nombre === agNom);
  return b * canjeF(ag ? ag.canje : 0);
}

function calculate() {
  if (order.size === 0) { alert('Agrega productos al pedido primero.'); return; }
  $('results-section').style.display = '';
  doCalc();
  $('results-section').scrollIntoView({behavior:'smooth'});
}

function doCalc() {
  const lines = [...order.values()].map(p => ({
    cod: p.c, desc: p.d.substring(0,55), cat: p.cn, uc: p.uc, qty: p.qty
  }));

  // Modo 1
  const m1 = AGENCIES
    .filter(a => AG_NAMES.includes(a.nombre))
    .map(ag => {
      let bru = 0, net = 0;
      const mis = [];
      for (const l of lines) {
        const b = getBruto(ag.nombre, l.uc);
        const n = getNeto(ag.nombre, l.uc);
        if (b === null) mis.push(l.cat);
        else { bru += b * l.qty; net += n * l.qty; }
      }
      return { ag, bru, net: mis.length ? null : net, mis };
    })
    .filter(r => r.net !== null)
    .sort((a,b) => a.net - b.net);

  // Modo 2
  let tot2 = 0;
  const m2 = lines.map(l => {
    let bestAg = null, bestN = Infinity, bestB = null;
    for (const ag of AGENCIES) {
      if (!AG_NAMES.includes(ag.nombre)) continue;
      const n = getNeto(ag.nombre, l.uc);
      if (n !== null && n < bestN) { bestN = n; bestB = getBruto(ag.nombre, l.uc); bestAg = ag; }
    }
    const sub = bestAg ? bestN * l.qty : null;
    if (sub !== null) tot2 += sub;
    return { ...l, bestAg, bestB, bestN, sub };
  });

  renderM1(m1);
  renderM2(m2, tot2);
  renderSummary(m1, tot2, m2);
}

// Re-calcular al cambiar canje
document.querySelectorAll('input[name="cj"]').forEach(r => {
  r.addEventListener('change', () => {
    $('margin-wrap').style.display = r.value === 'B' ? '' : 'none';
    if ($('results-section').style.display !== 'none') doCalc();
  });
});
$('margin-sl').addEventListener('input', () => {
  $('margin-lbl').textContent = $('margin-sl').value + '%';
  if ($('results-section').style.display !== 'none') doCalc();
});

function renderM1(rows) {
  const medals = ['&#127945;','&#127946;','&#127947;'];
  $('m1-body').innerHTML = rows.map((r, i) => {
    const cls = i < 3 ? 'rank-'+(i+1) : '';
    const isZero = r.net < 1;
    const netoHtml = isZero
      ? '<span class="neto0">$0 <small>(canje 100%)</small></span>'
      : '<span class="ptag">' + fmt(r.net) + '</span>';
    return '<tr class="' + cls + '">' +
      '<td>' + (medals[i] || (i+1)+'.') + '</td>' +
      '<td class="fw-semibold">' + esc(r.ag.nombre) + '</td>' +
      '<td class="text-end text-muted">' + fmt(r.bru) + '</td>' +
      '<td class="text-end">' + netoHtml + '</td>' +
      '<td><span class="badge bg-secondary">' + Math.round(r.ag.canje*100) + '%</span></td>' +
      '<td>' + esc(r.ag.freq) + '</td>' +
      '<td class="text-center"><span class="badge ' + (r.ag.dias>=5?'bg-success':'bg-warning text-dark') + '">' +
        r.ag.dias + 'd</span></td>' +
      '</tr>';
  }).join('');
}

function renderM2(rows, tot2) {
  $('m2-body').innerHTML = rows.map(r => {
    const isZero = r.sub !== null && r.sub < 1;
    const subHtml = isZero
      ? '<span class="neto0">$0</span>'
      : '<span class="ptag">' + fmt(r.sub) + '</span>';
    return '<tr>' +
      '<td><code>' + esc(r.cod) + '</code></td>' +
      '<td>' + esc(r.desc) + '</td>' +
      '<td><span class="bcat">' + esc(r.cat) + '</span></td>' +
      '<td class="text-center">' + r.qty + '</td>' +
      '<td class="fw-semibold">' + (r.bestAg ? esc(r.bestAg.nombre) : 'Sin precio') + '</td>' +
      '<td class="text-end">' + fmt(r.bestN) + '</td>' +
      '<td class="text-end">' + subHtml + '</td>' +
      '<td>' + (r.bestAg ? esc(r.bestAg.freq) : '') + '</td>' +
      '</tr>';
  }).join('');
  $('m2-total').innerHTML = '<span class="ptag">' + fmt(tot2) + '</span>';
}

function renderSummary(m1, tot2, m2) {
  if (!m1.length) { $('summary-cards').innerHTML = ''; return; }
  const best = m1[0];
  const diff = best.net - tot2;
  const pct  = best.net > 0 ? Math.abs(diff / best.net * 100).toFixed(1) : '0';
  const uniqAgs = [...new Set(m2.filter(r => r.bestAg).map(r => r.bestAg.nombre))];
  const pedInfo = [
    $('f-nped').value    ? 'Pedido #'+$('f-nped').value : '',
    $('f-cliente').value ? $('f-cliente').value : '',
    $('f-dest').value    ? $('f-dest').value : '',
  ].filter(Boolean).join(' | ');

  let diffBadge = '';
  if (Math.abs(diff) < 1) {
    diffBadge = '<span class="badge bg-secondary">Mismo costo</span>';
  } else if (diff > 0) {
    diffBadge = '<span class="badge bg-danger">Modo 1 cuesta ' + fmt(diff).replace('$','$') + ' mas</span>';
  } else {
    diffBadge = '<span class="badge bg-success">Modo 1 ahorra ' + fmt(-diff).replace('$','$') + '</span>';
  }

  $('summary-cards').innerHTML =
    '<div class="col-md-4">' +
    '<div class="card border-0 text-white" style="background:var(--azul)">' +
    '<div class="card-body text-center py-3">' +
    (pedInfo ? '<div class="small mb-1 opacity-75">' + esc(pedInfo) + '</div>' : '') +
    '<div class="small mb-1 opacity-75">Modo 1 &mdash; Mejor agencia unica</div>' +
    '<div class="fs-3 fw-bold">' + (best.net < 1 ? '$0 &#10003;' : fmt(best.net)) + '</div>' +
    '<div class="small mt-1 fw-semibold">' + esc(best.ag.nombre) + '</div>' +
    '<div class="small opacity-75">' + esc(best.ag.freq) + ' &middot; canje ' + Math.round(best.ag.canje*100) + '%</div>' +
    '</div></div></div>' +

    '<div class="col-md-4">' +
    '<div class="card border-0 text-white" style="background:#198754">' +
    '<div class="card-body text-center py-3">' +
    '<div class="small mb-1 opacity-75">Modo 2 &mdash; Mejor agencia por producto</div>' +
    '<div class="fs-3 fw-bold">' + (tot2 < 1 ? '$0 &#10003;' : fmt(tot2)) + '</div>' +
    '<div class="small mt-1">' + uniqAgs.length + ' agencia(s)</div>' +
    '<div class="small opacity-75">' + esc(uniqAgs.slice(0,2).join(', ')) + (uniqAgs.length>2?'...':'') + '</div>' +
    '</div></div></div>' +

    '<div class="col-md-4">' +
    '<div class="card border-0 bg-light">' +
    '<div class="card-body text-center py-3">' +
    '<div class="small text-muted mb-2">Comparativa</div>' +
    '<div class="mb-1">' + diffBadge + '</div>' +
    '<div class="small text-muted">' + pct + '% de diferencia</div>' +
    '<div class="small text-muted">' + (diff > 0 ? 'Conviene dividir el envio' : 'Conviene una sola agencia') + '</div>' +
    '</div></div></div>';
}

// ── Catalogo ─────────────────────────────────────────────────────────────────
const famSel = $('cat-fam');
[...new Set(CATALOG.map(p => p.f))].sort().forEach(f => {
  const o = document.createElement('option');
  o.value = f; o.textContent = f;
  famSel.appendChild(o);
});

function filterCat() {
  const q   = $('cat-q').value.toLowerCase();
  const fam = $('cat-fam').value;
  catFiltered = CATALOG.filter(p =>
    (!fam || p.f === fam) &&
    (!q   || p.c.toLowerCase().includes(q) ||
             p.d.toLowerCase().includes(q) ||
             p.r.toLowerCase().includes(q))
  );
  catPage = 1;
  renderCat();
}

function renderCat() {
  const total = catFiltered.length;
  const pages = Math.max(1, Math.ceil(total / PG));
  const start = (catPage - 1) * PG;
  const slice = catFiltered.slice(start, start + PG);

  $('cat-cnt').textContent = total.toLocaleString() + ' productos' +
    (total < CATALOG.length ? ' (filtrado)' : '');
  $('cat-pg-info').textContent = 'Mostrando ' + (start+1) + '-' +
    Math.min(start+PG, total) + ' de ' + total;

  $('cat-body').innerHTML = slice.map(p => {
    const inOrd = order.has(p.c);
    const stCls = p.st > 50 ? 'text-success' : p.st > 0 ? 'text-warning' : 'text-danger';
    const mkCls = p.ms < 2 && p.ms > 0 ? 'text-danger fw-semibold' : '';
    return '<tr class="cat-tr" data-cod="' + esc(p.c) + '">' +
      '<td><code class="text-primary">' + esc(p.c) + '</code></td>' +
      '<td>' + esc(p.d.substring(0,60)) + '</td>' +
      '<td class="text-muted small">' + esc(p.f) + '</td>' +
      '<td class="text-muted small">' + esc((p.r||'').substring(0,28)) + '</td>' +
      '<td class="text-end">' + (p.pl > 0 ? '$'+p.pl.toLocaleString('es-UY') : '-') + '</td>' +
      '<td class="text-end ' + stCls + ' fw-semibold">' + p.st.toLocaleString() + '</td>' +
      '<td class="text-end">' + p.d3.toLocaleString() + '</td>' +
      '<td class="text-end ' + mkCls + '">' + (p.ms > 0 ? p.ms : '-') + '</td>' +
      '<td><span class="bcat">' + esc(p.cn) + '</span></td>' +
      '<td>' +
        '<button class="btn btn-sm ' + (inOrd?'btn-success':'btn-outline-primary') + ' py-0 add-cat"' +
        ' data-cod="' + esc(p.c) + '">' +
        '<i class="bi ' + (inOrd?'bi-cart-check':'bi-cart-plus') + '"></i></button>' +
      '</td></tr>';
  }).join('');

  // listeners
  $('cat-body').querySelectorAll('.cat-tr').forEach(tr => {
    tr.addEventListener('click', e => {
      if (!e.target.closest('.add-cat')) showDetail(tr.dataset.cod);
    });
  });
  $('cat-body').querySelectorAll('.add-cat').forEach(btn => {
    btn.addEventListener('click', e => {
      e.stopPropagation();
      const p = CATALOG.find(x => x.c === btn.dataset.cod);
      if (p) { addProduct(p); showTab('calc'); }
    });
  });

  // pagination
  const pgDiv = $('cat-pg');
  if (pages <= 1) { pgDiv.innerHTML = ''; return; }
  const nums = [...new Set([1,
    ...Array.from({length:5}, (_,i) => Math.max(1,Math.min(pages, catPage-2+i))),
    pages])].sort((a,b)=>a-b);
  let pgH = '';
  nums.forEach((n,i) => {
    if (i > 0 && nums[i] - nums[i-1] > 1) pgH += '<span class="px-1">...</span>';
    pgH += '<span class="pg-btn ' + (n===catPage?'active':'') + '" data-pg="'+n+'">'+n+'</span>';
  });
  pgDiv.innerHTML =
    '<span class="pg-btn" data-pg="'+(catPage-1)+'"><i class="bi bi-chevron-left"></i></span>' +
    pgH +
    '<span class="pg-btn" data-pg="'+(catPage+1)+'"><i class="bi bi-chevron-right"></i></span>';
  pgDiv.querySelectorAll('.pg-btn').forEach(btn => {
    btn.addEventListener('click', () => {
      const p = parseInt(btn.dataset.pg);
      if (p >= 1 && p <= pages) { catPage = p; renderCat(); $('cat-body').closest('.card').scrollIntoView({behavior:'smooth'}); }
    });
  });
}

$('cat-q').addEventListener('input', filterCat);
$('cat-fam').addEventListener('change', filterCat);

// ── Detail modal ─────────────────────────────────────────────────────────────
function showDetail(cod) {
  const p = CATALOG.find(x => x.c === cod);
  if (!p) return;
  const rows = AGENCIES
    .filter(a => AG_NAMES.includes(a.nombre))
    .map(ag => {
      const b = getBruto(ag.nombre, p.uc);
      const n = getNeto(ag.nombre, p.uc);
      return b !== null ? {ag, b, n} : null;
    })
    .filter(Boolean)
    .sort((a,b) => a.n - b.n);

  const prev = document.getElementById('detail-modal');
  if (prev) prev.remove();

  const el = document.createElement('div');
  el.id = 'detail-modal';
  el.innerHTML =
    '<div class="modal fade" id="dm-inner" tabindex="-1">' +
    '<div class="modal-dialog modal-lg">' +
    '<div class="modal-content">' +
    '<div class="modal-header" style="background:var(--azul);color:#fff">' +
    '<h5 class="modal-title"><code>' + esc(p.c) + '</code> &mdash; ' + esc(p.d.substring(0,55)) + '</h5>' +
    '<button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>' +
    '</div>' +
    '<div class="modal-body">' +
    '<div class="row mb-3">' +
    '<div class="col-sm-4"><small class="text-muted">Familia</small><br>' + esc(p.f) + '</div>' +
    '<div class="col-sm-4"><small class="text-muted">Ramo</small><br>' + esc(p.r||'-') + '</div>' +
    '<div class="col-sm-4"><small class="text-muted">Categoria envio</small><br><span class="bcat">' + esc(p.cn) + '</span></div>' +
    '</div>' +
    '<div class="row mb-3 text-center">' +
    '<div class="col-3"><small class="text-muted d-block">P. Lista</small><strong>' + (p.pl > 0 ? '$'+p.pl.toLocaleString() : '-') + '</strong></div>' +
    '<div class="col-3"><small class="text-muted d-block">Stock</small><strong class="' + (p.st>0?'text-success':'text-danger') + '">' + p.st.toLocaleString() + '</strong></div>' +
    '<div class="col-3"><small class="text-muted d-block">Pedido</small><strong>' + p.pe.toLocaleString() + '</strong></div>' +
    '<div class="col-3"><small class="text-muted d-block">Disp. 30d</small><strong>' + p.d3.toLocaleString() + '</strong></div>' +
    '</div>' +
    '<h6>Precios de envio (por unidad, neto)</h6>' +
    '<table class="table table-sm"><thead class="table-light"><tr><th>#</th><th>Agencia</th><th class="text-end">Bruto</th><th class="text-end">Neto</th><th>Canje</th><th>Frec.</th></tr></thead><tbody>' +
    rows.map((r,i) =>
      '<tr class="' + (i<3?'rank-'+(i+1):'') + '">' +
      '<td>' + (['&#127945;','&#127946;','&#127947;'][i]||i+1+'.') + '</td>' +
      '<td class="fw-semibold">' + esc(r.ag.nombre) + '</td>' +
      '<td class="text-end text-muted">' + fmt(r.b) + '</td>' +
      '<td class="text-end">' + (r.n < 1 ? '<span class="neto0">$0</span>' : '<span class="ptag">'+fmt(r.n)+'</span>') + '</td>' +
      '<td><span class="badge bg-secondary">' + Math.round(r.ag.canje*100) + '%</span></td>' +
      '<td>' + esc(r.ag.freq) + '</td>' +
      '</tr>'
    ).join('') +
    '</tbody></table>' +
    '</div>' +
    '<div class="modal-footer">' +
    '<button class="btn btn-p" id="dm-add">'+
    '<i class="bi bi-cart-plus me-1"></i>Agregar al pedido</button>' +
    '<button class="btn btn-secondary" data-bs-dismiss="modal">Cerrar</button>' +
    '</div></div></div></div>';

  document.body.appendChild(el);
  const modal = new bootstrap.Modal(document.getElementById('dm-inner'));
  modal.show();
  document.getElementById('dm-add').addEventListener('click', () => {
    addProduct(p);
    modal.hide();
    showTab('calc');
  });
}

// ── Init ─────────────────────────────────────────────────────────────────────
$('f-fecha').value = today();
filterCat();
</script>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>'''

# ── Importar template actualizado (responsive + Supabase + todas las features) ──
from html_template import HTML_TEMPLATE

# ── Credenciales Supabase (desde .env o variables de entorno) ──────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import os
SUPABASE_URL = os.getenv('SUPABASE_URL', '')
SUPABASE_KEY = os.getenv('SUPABASE_KEY', '')

# ── Substituir placeholders con JSON real ─────────────────────────────────────
HTML = (HTML_TEMPLATE
    .replace('__CATALOG__',      catalog_json)
    .replace('__AGENCIES__',     agencies_json)
    .replace('__PRICES__',       prices_json)
    .replace('__UCATS__',        ucats_json)
    .replace('__AG_NAMES__',     ag_names_json)
    .replace('__SUPABASE_URL__', SUPABASE_URL)
    .replace('__SUPABASE_KEY__', SUPABASE_KEY)
)

out = Path('petinsa_envios.html')
out.write_text(HTML, encoding='utf-8')
print(f"\nArchivo generado: {out}  ({out.stat().st_size // 1024} KB)")
print("Abrir en el navegador.")
