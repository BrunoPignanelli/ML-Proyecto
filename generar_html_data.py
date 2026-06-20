"""
Genera petinsa_envios.html — app standalone para PETINSA.
Ejecutar: python generar_html.py
"""
import openpyxl, re, unicodedata, json
from pathlib import Path
import copy

# ═══════════════════════════════════════════════════════
# 1. FUNCIONES DE CLASIFICACIÓN
# ═══════════════════════════════════════════════════════

def norm(s):
    if s is None: return ''
    s = str(s).replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def extract_rim(desc):
    d = str(desc).upper()
    m = re.search(r'\d\s+-\s+(\d{2,3}(?:\.\d)?)\b', d)
    if m: return float(m.group(1))
    m = re.search(r'R\s*(\d{2}(?:\.\d)?)\b', d)
    if m: return float(m.group(1))
    m = re.search(r'(?<=[\d/\s])-(\d{2,3}(?:\.\d)?)\b', d)
    if m: return float(m.group(1))
    return None

def extract_amp(desc):
    m = re.search(r'\b(\d{2,3})A\b', str(desc).upper())
    return int(m.group(1)) if m else None

UNIFIED_CATS = {
    'cub_moto':          'Cubierta Moto',
    'cub_auto_r12_r14':  'Cubierta Auto R12-R14',
    'cub_auto_r15_r19':  'Cubierta Auto/Camioneta R15-R19',
    'cub_camion_r20_r22':'Cubierta Camion R20-R22.5',
    'cub_agro_del':      'Cubierta Agricola Delantera',
    'cub_agro_tras_med': 'Cubierta Agricola Trasera Med. 24-26"',
    'cub_agro_tras_gde': 'Cubierta Agricola Trasera Gde. 28-36"',
    'cub_agro_tras_xgde':'Cubierta Agricola/Vial +38"',
    'camara':            'Camara',
    'bateria_chica':     'Bateria Chica (<=110 Ah)',
    'bateria_grande':    'Bateria Grande (>110 Ah)',
    'lubricante_caja':   'Lubricante en caja/lata',
    'lubricante_tambor': 'Lubricante Tambor 205L',
    'bulto_general':     'Bulto General',
}

def classify(familia, desc):
    f = str(familia).upper()
    d = str(desc).upper()
    if 'LUBRICANTE' in f:
        return 'lubricante_tambor' if ('TAMBOR' in d or '205' in d) else 'lubricante_caja'
    if 'BATERIA' in f:
        amp = extract_amp(desc)
        return 'bateria_grande' if amp and amp > 110 else 'bateria_chica'
    if 'ACCES' in f or 'CAM MOTO' in f:
        return 'camara'
    if 'MOTO' in f and 'CAM' not in f:
        return 'cub_moto'
    if 'GIGANTE' in f:
        return 'cub_camion_r20_r22'
    if 'VIAL' in f or 'IND.' in f:
        rim = extract_rim(desc)
        return 'cub_agro_tras_xgde' if rim and rim >= 38 else 'cub_agro_tras_gde'
    if 'AGR DEL' in f or 'AGRICOLA DEL' in f:
        rim = extract_rim(desc)
        if rim:
            if rim >= 38: return 'cub_agro_tras_xgde'
            if rim >= 28: return 'cub_agro_tras_gde'
            if rim >= 24: return 'cub_agro_tras_med'
        return 'cub_agro_del'
    if 'AGR TRAS' in f or 'AGRICOLA TRAS' in f:
        rim = extract_rim(desc)
        if rim:
            if rim >= 38: return 'cub_agro_tras_xgde'
            if rim >= 30: return 'cub_agro_tras_gde'
            if rim >= 24: return 'cub_agro_tras_med'
        return 'cub_agro_tras_med'
    if 'PASEO' in f:
        rim = extract_rim(desc)
        if rim:
            if rim <= 14: return 'cub_auto_r12_r14'
            if rim <= 19: return 'cub_auto_r15_r19'
            return 'cub_camion_r20_r22'
        return 'cub_auto_r15_r19'
    if 'CAMIONETA' in f or 'PICK UP' in f:
        rim = extract_rim(desc)
        if rim:
            if rim <= 14: return 'cub_auto_r12_r14'
            if rim <= 19: return 'cub_auto_r15_r19'
            return 'cub_camion_r20_r22'
        return 'cub_auto_r15_r19'
    return 'bulto_general'

# ═══════════════════════════════════════════════════════
# 2. MAPEO: categoría unificada → label de la agencia
# ═══════════════════════════════════════════════════════

MAPPING = {
    'DAC': {
        'cub_moto':'Atado moto','cub_auto_r12_r14':'Rodado 12,13,14',
        'cub_auto_r15_r19':'Rodado 15 a 18','cub_camion_r20_r22':'Rodado 20 a 22,5',
        'cub_agro_del':'Cubierta tractor AGRO 24 a 42','cub_agro_tras_med':'Cubierta tractor AGRO 24 a 42',
        'cub_agro_tras_gde':'Cubierta tractor AGRO 24 a 42','cub_agro_tras_xgde':'Cubierta tractor AGRO 24 a 42',
        'camara':'Bulto hasta 15kg','bateria_chica':'Bateria chica','bateria_grande':'Bateria grande',
        'lubricante_caja':'Bulto hasta 15kg','lubricante_tambor':'Tambor 205lts','bulto_general':'Bulto hasta 15kg',
    },
    'NASAZZI': {
        'cub_moto':'Rodado 15 a 19','cub_auto_r12_r14':'Rodado 13 y 14',
        'cub_auto_r15_r19':'Rodado 15 a 19','cub_camion_r20_r22':'Rodado 20 a 22,5',
        'cub_agro_del':'Rodado 24 y 25','cub_agro_tras_med':'Rodado 24 y 25',
        'cub_agro_tras_gde':'Rodado 26 a 32','cub_agro_tras_xgde':'Cub viales de 38 a 46 extra grande',
        'camara':'Bultos','bateria_chica':'Bat hasta 110 amp','bateria_grande':'Bat mayor 110 amp',
        'lubricante_caja':'Bultos lubricantes','lubricante_tambor':'Tambor 205lts','bulto_general':'Bultos',
    },
    'MEGAM': {
        'cub_moto':'Bat y bultos','cub_auto_r12_r14':'Cub 13 y 14',
        'cub_auto_r15_r19':'Cubierta auto (15 a 17)','cub_camion_r20_r22':'Cubierta camion',
        'cub_agro_del':'Tractor chico','cub_agro_tras_med':'Tractor chico',
        'cub_agro_tras_gde':'Tractor grande','cub_agro_tras_xgde':'Tractor grande',
        'camara':'Bat y bultos','bateria_chica':'Bat y bultos','bateria_grande':'Bat y bultos',
        'lubricante_caja':'Bat y bultos','lubricante_tambor':'Tambor 205lts','bulto_general':'Bat y bultos',
    },
    'Transportes Nagar sas (Expreso Ruta 1)': {
        'cub_moto':'Auto','cub_auto_r12_r14':'Auto','cub_auto_r15_r19':'Camioneta',
        'cub_camion_r20_r22':'Camion','cub_agro_del':'Tractor del y 17,5',
        'cub_agro_tras_med':'Tractor traseras','cub_agro_tras_gde':'Tractor grande',
        'cub_agro_tras_xgde':'Tractor grande','camara':'Bultos','bateria_chica':'Baterias',
        'bateria_grande':'Baterias','lubricante_caja':'Bultos','lubricante_tambor':'Tambor','bulto_general':'Bultos',
    },
    'SELEGUIN': {
        'cub_moto':'Bultos, atados, bolsas, cajas','cub_auto_r12_r14':'Cubierta auto',
        'cub_auto_r15_r19':'Cubierta camioneta','cub_camion_r20_r22':'Cubierta camion',
        'cub_agro_del':'Bultos, atados, bolsas, cajas','cub_agro_tras_med':'Tractor mediana 24 a 26',
        'cub_agro_tras_gde':'Tractor grande 28 a 30','cub_agro_tras_xgde':'Tractor extra grande +30',
        'camara':'Bultos, atados, bolsas, cajas','bateria_chica':'Baterias chicas','bateria_grande':'Baterias grandes',
        'lubricante_caja':'Bultos, atados, bolsas, cajas','lubricante_tambor':None,'bulto_general':'Bultos, atados, bolsas, cajas',
    },
    'EXPRESO ROCHA': {
        'cub_moto':'Bultos camaras','cub_auto_r12_r14':'Cubierta chica hasta 205/70.16',
        'cub_auto_r15_r19':'Cubierta grande hasta 750.16','cub_camion_r20_r22':'Cubierta camion',
        'cub_agro_del':'Neumatico agricolas y viales hasta 18.4/15-26',
        'cub_agro_tras_med':'Neum. Agricolas y viales mas de 23.1/18-26',
        'cub_agro_tras_gde':'Neum. Agricolas y viales mas de 23.1/18-26',
        'cub_agro_tras_xgde':'Neum. Agricolas y viales mas de 23.1/18-26',
        'camara':'Bultos camaras','bateria_chica':'Baterias','bateria_grande':'Baterias',
        'lubricante_caja':'Bultos camaras','lubricante_tambor':None,'bulto_general':'Bultos camaras',
    },
    'BULEVAR (ACC)': {
        'cub_moto':'Atado cub moto','cub_auto_r12_r14':'Cub auto  hasta rod 14',
        'cub_auto_r15_r19':'Cub camioneta 15 a 19','cub_camion_r20_r22':'Cub camion',
        'cub_agro_del':'Cub tractor chica (hasta 22)','cub_agro_tras_med':'Cub tractor mediana (24)',
        'cub_agro_tras_gde':'Cub tractor grande (de 25 a 36)','cub_agro_tras_xgde':'Cub vial rod 38 a 42',
        'camara':'Bolsas','bateria_chica':'Bat chicas','bateria_grande':'Bat grandes',
        'lubricante_caja':'Caja aceite 1 y 4lts','lubricante_tambor':'Tambor 205lts','bulto_general':'Bolsas',
    },
    'PERICO': {
        'cub_moto':'Atado cub moto','cub_auto_r12_r14':'Cub auto  hasta rod 14',
        'cub_auto_r15_r19':'Cub camioneta 15 a 19','cub_camion_r20_r22':'Cub camion',
        'cub_agro_del':'Cub tractor chica','cub_agro_tras_med':'Cub tractor mediana',
        'cub_agro_tras_gde':'Cub tractor grande','cub_agro_tras_xgde':'Cub vial rod 38 a 42',
        'camara':'Bolsas','bateria_chica':'Bat chicas','bateria_grande':'Bat grandes',
        'lubricante_caja':'Caja aceite 1 y 4lts','lubricante_tambor':'Tambor 205lts','bulto_general':'Bolsas',
    },
    'TRUJILLO': {
        'cub_moto':'Bultos','cub_auto_r12_r14':'Paseo hasta 16','cub_auto_r15_r19':'Paseo hasta 16',
        'cub_camion_r20_r22':'Camion','cub_agro_del':'Agricola delantera',
        'cub_agro_tras_med':'Agricola mediana','cub_agro_tras_gde':'Agricola extra grande',
        'cub_agro_tras_xgde':'Agricola extra grande','camara':'Bultos',
        'bateria_chica':'Bat hasta 150 amp','bateria_grande':'Bat + 150 amp',
        'lubricante_caja':'Cajas lubricantes','lubricante_tambor':'Tambor 205lts','bulto_general':'Bultos',
    },
    'GONFER': {
        'cub_moto':'Bultos','cub_auto_r12_r14':'Cubierta de auto',
        'cub_auto_r15_r19':'Cubierta de camioneta','cub_camion_r20_r22':'Cubierta de camion chico',
        'cub_agro_del':'Agricola chica','cub_agro_tras_med':'Agricola hasta rod 28',
        'cub_agro_tras_gde':'Agricola grande rod 30 en adelante','cub_agro_tras_xgde':'Agricola grande rod 30 en adelante',
        'camara':'Bultos','bateria_chica':'Bat hasta 110 amp','bateria_grande':'Bat mayor 110 amp',
        'lubricante_caja':'Bultos lubricantes','lubricante_tambor':'Tambor 205lts','bulto_general':'Bultos',
    },
    '3EME (El Chambon)': {
        'cub_moto':'Atado cub moto','cub_auto_r12_r14':'Cub auto  hasta rod 14',
        'cub_auto_r15_r19':'Cub camioneta 15 a 19','cub_camion_r20_r22':'Cub camion',
        'cub_agro_del':'Cub tractor chica','cub_agro_tras_med':'Cub tractor mediana',
        'cub_agro_tras_gde':'Cub tractor grande','cub_agro_tras_xgde':'Cub vial rod 38 a 42',
        'camara':'Bolsas','bateria_chica':'Bat chicas','bateria_grande':'Bat grandes',
        'lubricante_caja':'Caja aceite 1 y 4lts','lubricante_tambor':'Tambor 205lts','bulto_general':'Bolsas',
    },
    'Martin Escudero (Lascano)': {
        'cub_moto':'Auto','cub_auto_r12_r14':'Auto','cub_auto_r15_r19':'Camioneta',
        'cub_camion_r20_r22':'Camion','cub_agro_del':'Agricola chica',
        'cub_agro_tras_med':'Agricola mediana','cub_agro_tras_gde':'Agricola grande',
        'cub_agro_tras_xgde':'Agricola grande','camara':'Bultos/atados',
        'bateria_chica':'Baterias chicas','bateria_grande':'Baterias grandes',
        'lubricante_caja':'Bultos lubricantes','lubricante_tambor':'Tambor aceite','bulto_general':'Bultos/atados',
    },
    'Arzuaga': {
        'cub_moto':'Atado cubiertas moto','cub_auto_r12_r14':'cubiertas r13 r14',
        'cub_auto_r15_r19':'cubiertas r16 r18','cub_camion_r20_r22':'Cubiertas camion',
        'cub_agro_del':'Cubiertas tractor chica','cub_agro_tras_med':'Cubiertas tractor mediana',
        'cub_agro_tras_gde':'Cubiertas tractor grande','cub_agro_tras_xgde':'Cubiertas tractor grande',
        'camara':'Paquetes/cajas','bateria_chica':'Baterias chicas','bateria_grande':'Baterias camion',
        'lubricante_caja':'Paquetes/cajas','lubricante_tambor':'Tanque aceite','bulto_general':'Paquetes/cajas',
    },
    'Franchi': {
        'cub_moto':'Bultos','cub_auto_r12_r14':'Cub.auto','cub_auto_r15_r19':'Cub.camioneta',
        'cub_camion_r20_r22':'Cub.camion','cub_agro_del':'Bultos',
        'cub_agro_tras_med':'Cub. agricola mediana','cub_agro_tras_gde':'Cub. agricola gde',
        'cub_agro_tras_xgde':'Cub. agricola gde','camara':'Bultos',
        'bateria_chica':'Baterias chicas','bateria_grande':'Baterias grandes',
        'lubricante_caja':'Bultos','lubricante_tambor':None,'bulto_general':'Bultos',
    },
}

# ═══════════════════════════════════════════════════════
# 3. CARGA DE DATOS
# ═══════════════════════════════════════════════════════

print("Cargando Libro1.xlsx...")
wb1 = openpyxl.load_workbook('Libro1.xlsx')
ws1 = wb1['Hoja1']
rows1 = list(ws1.iter_rows(values_only=True))
cols1 = [str(c).strip() if c else '' for c in rows1[0]]

catalog = []
for row in rows1[1:]:
    if not any(row): continue
    r = dict(zip(cols1, row))
    fam  = str(r.get('Familia', '') or '').strip()
    desc = str(r.get('Descripción de Producto', '') or '').strip()
    cod_raw = r.get('Prod.')
    cod = str(cod_raw).strip().rstrip('.0') if cod_raw is not None else ''
    if not cod or not desc: continue
    ucat = classify(fam, desc)
    def safe_int(v):
        try: return int(float(v or 0))
        except: return 0
    def safe_float(v):
        try: return round(float(v or 0), 2)
        except: return 0.0
    catalog.append({
        'c': cod, 'd': desc, 'f': fam,
        'r': str(r.get('Ramo', '') or '').strip(),
        'pl': safe_float(r.get('P. Lista')),
        'st': safe_int(r.get('Stock')),
        'pe': safe_int(r.get('Pedido')),
        'd3': safe_int(r.get('Disp 30d')),
        'ms': safe_float(r.get('Meses Stk')),
        'uc': ucat,
        'cn': UNIFIED_CATS[ucat],
    })

print(f"  {len(catalog)} productos cargados.")

print("Cargando COMPARATIVA AGENCIAS.xlsx (Mayo 2026)...")
wb2 = openpyxl.load_workbook('COMPARATIVA AGENCIAS.xlsx', data_only=True)
ws2 = wb2['Costo ag. Mayo 2026']
ag_rows = list(ws2.iter_rows(values_only=True))
agency_name_row = ag_rows[5]
canje_row       = ag_rows[2]
freq_row        = ag_rows[4]
data_rows       = ag_rows[7:]

ncols = len(agency_name_row)
acg = {}
col = 0
while col < ncols:
    name = agency_name_row[col]
    if name and str(name).strip():
        nxt = col + 1
        while nxt < ncols and (agency_name_row[nxt] is None or str(agency_name_row[nxt]).strip() == ''):
            nxt += 1
        acg[str(name).strip()] = list(range(col, nxt))
        col = nxt
    else:
        col += 1

IVA = 1.22

def pick_price_col(ag, ci):
    pc = ci[1] if len(ci) > 1 else ci[0]
    if 'GONFER' in ag.upper() and len(ci) >= 4: pc = ci[3]
    return ci[0], pc

def parse_canje(val):
    if isinstance(val, (int, float)): return min(float(val), 1.0)
    return 0.0

def parse_freq(text):
    if not text: return ('L a V', 5)
    t = str(text).lower()
    if 'cuando tiene carga' in t: return ('Bajo pedido', 1)
    if 'mart y jue' in t or 'martes y jueves' in t: return ('Mar y Jue', 2)
    if 'l, mier y v' in t: return ('L, Mier y V', 3)
    return ('L a V', 5)

agencies_meta  = {}
agency_prices  = {}

for ag, ci in acg.items():
    cat_col, price_col = pick_price_col(ag, ci)
    prices = {}
    for row in data_rows:
        cat   = row[cat_col]   if cat_col   < len(row) else None
        price = row[price_col] if price_col < len(row) else None
        if cat and isinstance(cat, str) and cat.strip() and price and isinstance(price, (int, float)):
            prices[norm(cat.strip())] = float(price)
    agency_prices[ag] = prices
    canje_col = ci[2] if ('GONFER' in ag.upper() and len(ci) >= 3) else price_col
    raw_canje = canje_row[canje_col] if canje_col < len(canje_row) else None
    raw_freq  = freq_row[cat_col]    if cat_col   < len(freq_row)  else None
    freq_label, freq_days = parse_freq(raw_freq)
    agencies_meta[ag] = {
        'nombre': ag,
        'canje':  parse_canje(raw_canje),
        'freq':   freq_label,
        'dias':   freq_days,
        'iva':    'sin IVA' if 'GONFER' in ag.upper() else 'IVA incl.',
    }

def build_prices(mapping):
    pj = {}
    for ag, ag_map in mapping.items():
        if ag not in agencies_meta: continue
        pj[ag] = {}
        for ucat, label in ag_map.items():
            if label is None: continue
            bruto = agency_prices.get(ag, {}).get(norm(label))
            if bruto:
                if agencies_meta[ag]['iva'] == 'sin IVA':
                    bruto = round(bruto * IVA, 2)
                pj[ag][ucat] = bruto
    return pj

prices_data = build_prices(MAPPING)
print(f"  {len(agencies_meta)} agencias procesadas.")

# ═══════════════════════════════════════════════════════
# 4. SERIALIZAR JSON
# ═══════════════════════════════════════════════════════
catalog_json   = json.dumps(catalog,                      ensure_ascii=False, separators=(',',':'))
agencies_json  = json.dumps(list(agencies_meta.values()), ensure_ascii=False, separators=(',',':'))
prices_json    = json.dumps(prices_data,                  ensure_ascii=False, separators=(',',':'))
ucats_json     = json.dumps(UNIFIED_CATS,                 ensure_ascii=False, separators=(',',':'))
ag_names_json  = json.dumps(list(MAPPING.keys()),         ensure_ascii=False, separators=(',',':'))
